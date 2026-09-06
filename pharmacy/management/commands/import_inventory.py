from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

import openpyxl
from django.conf import settings
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from pharmacy.models import Category, Stock


class Command(BaseCommand):
	help = "Replace Stock with records from the workbook's Master Inventory sheet."

	def add_arguments(self, parser):
		parser.add_argument(
			"--file",
			default=str(Path(settings.BASE_DIR) / "Pharmacy inventory book.xlsx"),
			help="Path to the Excel workbook (defaults to the project workbook).",
		)
		parser.add_argument(
			"--sheet",
			default="Master Inventory",
			help="Worksheet containing the master inventory.",
		)
		parser.add_argument(
			"--only-if-empty",
			action="store_true",
			help="Skip the import when inventory records already exist.",
		)

	def handle(self, *args, **options):
		if options["only_if_empty"] and Stock.objects.exists():
			self.stdout.write("Inventory already contains records; skipping workbook import.")
			return

		workbook_path = Path(options["file"])
		if not workbook_path.exists():
			raise CommandError(f"Workbook not found: {workbook_path}")

		try:
			workbook = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
		except (OSError, ValueError) as error:
			raise CommandError(f"Could not open workbook: {error}") from error

		sheet_name = options["sheet"]
		if sheet_name not in workbook.sheetnames:
			raise CommandError(f"Worksheet not found: {sheet_name}")

		sheet = workbook[sheet_name]
		rows = sheet.iter_rows(values_only=True)
		header_row = self._find_header(rows)
		headers = {self._normalise(value): index for index, value in enumerate(header_row) if value}
		required = {"itemid", "itemname", "currentstockqty", "buyingpricekes", "sellingpricekes"}
		missing = required - headers.keys()
		if missing:
			raise CommandError(f"Missing required columns: {', '.join(sorted(missing))}")

		imported = []
		skipped = []
		for row_number, row in enumerate(rows, start=6):
			item_id = self._text(self._value(row, headers, "itemid"))
			item_name = self._text(self._value(row, headers, "itemname"))
			if not item_id or not item_name or item_name.lower().startswith("total valuation"):
				skipped.append(row_number)
				continue

			imported.append({
				"item_id": item_id,
				"item_name": item_name,
				"category": self._text(self._value(row, headers, "category")),
				"quantity": self._integer(self._value(row, headers, "currentstockqty"), row_number, "Current Stock (Qty)"),
				"buying_price": self._decimal(self._value(row, headers, "buyingpricekes"), row_number, "Buying Price (KES)"),
				"selling_price": self._decimal(self._value(row, headers, "sellingpricekes"), row_number, "Selling Price (KES)"),
				"supplier": self._text(self._value(row, headers, "supplier")),
				"valid_to": self._date(self._value(row, headers, "expirydate")),
				"reorder_level": self._integer(self._value(row, headers, "minstocklevel"), row_number, "Min Stock Level"),
			})

		item_ids = [item["item_id"] for item in imported]
		if len(item_ids) != len(set(item_ids)):
			raise CommandError("The workbook contains duplicate Item ID values.")

		with transaction.atomic():
			Stock.objects.all().delete()
			categories = {}
			for item in imported:
				category_name = item.pop("category")
				category = None
				if category_name:
					category = categories.get(category_name.casefold())
					if category is None:
						category, _ = Category.objects.get_or_create(name=category_name)
						categories[category_name.casefold()] = category
				Stock.objects.create(category=category, drug_name=item.pop("item_name"), **item)

		self.stdout.write(self.style.SUCCESS(
			f"Imported {len(imported)} inventory items from {workbook_path.name}; skipped {len(skipped)} rows."
		))
		if skipped:
			self.stdout.write(f"Skipped worksheet rows: {', '.join(map(str, skipped))}")

	@staticmethod
	def _find_header(rows):
		for row in rows:
			if "itemid" in {Command._normalise(value) for value in row if value}:
				return row
		raise CommandError("Could not find the inventory header row.")

	@staticmethod
	def _normalise(value):
		return "".join(str(value).lower().split()).replace("(", "").replace(")", "")

	@staticmethod
	def _value(row, headers, name):
		index = headers.get(name)
		return row[index] if index is not None and index < len(row) else None

	@staticmethod
	def _text(value):
		return str(value).strip() if value is not None and str(value).strip() else None

	@staticmethod
	def _integer(value, row_number, column):
		if value in (None, ""):
			return 0
		try:
			return int(float(str(value).replace(",", "")))
		except (TypeError, ValueError) as error:
			raise CommandError(f"Invalid {column} on worksheet row {row_number}: {value}") from error

	@staticmethod
	def _decimal(value, row_number, column):
		if value in (None, ""):
			return Decimal("0.00")
		try:
			return Decimal(str(value).replace(",", "")).quantize(Decimal("0.01"))
		except (InvalidOperation, ValueError) as error:
			raise CommandError(f"Invalid {column} on worksheet row {row_number}: {value}") from error

	@staticmethod
	def _date(value):
		if isinstance(value, datetime):
			return value.date()
		if isinstance(value, date):
			return value
		if not value:
			return None
		for date_format in ("%d/%m/%Y", "%Y-%m-%d", "%m/%d/%Y"):
			try:
				return datetime.strptime(str(value).strip(), date_format).date()
			except ValueError:
				continue
		raise CommandError(f"Invalid Expiry Date: {value}")
